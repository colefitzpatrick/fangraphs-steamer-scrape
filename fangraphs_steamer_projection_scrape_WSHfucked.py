from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import time
import openpyxl
import requests
import warnings
os.chdir('c:\\Python\\colefitzpatrick_python\\FantasyBaseballProjection')

url = "https://www.fangraphs.com/projections.aspx?pos=all&stats=bat&type=steamer"
url2 = "https://www.fangraphs.com/projections.aspx?pos=all&stats=pit&type=steamer&team=0&lg=all&players=0"

wb_write = openpyxl.load_workbook('steamer.xlsx')
ws_write = wb_write["hitters"]
ws2_write = wb_write["pitchers"]
#wb3 = openpyxl.load_workbook('fantrax.xlsx')
#ws3 = wb3["Sheet1"]
wb4 = openpyxl.load_workbook('teamacronyms.xlsx')
ws4 = wb4["Sheet1"]
fantrax_wb = openpyxl.load_workbook('fantrax.xlsx')
fantrax_ws = fantrax_wb["Sheet1"]
standings_wb = openpyxl.load_workbook('standings.xlsx')
standings_ws = standings_wb["Sheet1"]

def fantraxscrape():
    fantraxurl = "https://www.fantrax.com/login"
    fantraxurl2 = "https://www.fantrax.com/fantasy/league/v2omq3g1k5phrpoi/home"
    fantraxurl3 = "https://www.fantrax.com/fantasy/league/v2omq3g1k5phrpoi/players"


    # create a new Chrome session
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(chrome_options=options)
    #driver = webdriver.Firefox()
    time.sleep(5)
    driver.get(fantraxurl)
    time.sleep(5)

    #enters login information
    #username = driver.find_element_by_id("mat-input-0")
    username = driver.find_element_by_xpath("//input[@formcontrolname='email']")
    username.clear()
    username.send_keys("OMITTED")            #<-------username input

    #password = driver.find_element_by_id("mat-input-1")
    password = driver.find_element_by_xpath("//input[@formcontrolname='password']")
    password.clear()
    password.send_keys("OMITTED")             #<-------password input

    driver.implicitly_wait(3)

    #clicks submit
    loginbutton = driver.find_element_by_xpath("//button[@type='submit']")
    driver.execute_script("arguments[0].click();", loginbutton)

    time.sleep(5)
    driver.get(fantraxurl2)
    time.sleep(3)

    soup_level3=BeautifulSoup(driver.page_source, 'lxml')

    standingsrow = 1
    
    #scrapes the standings on the league homepage
    for tr in soup_level3.findAll("tr", {"class": "ng-star-inserted"}):
        info = []
        for td in tr.findAll("td", {"class": "ng-star-inserted"}):
            info.append(td.text)
        if len(info) > 0:
            teamname = info[1].strip()
            teamrecord = info[2].strip()
            dashpos = teamrecord.find('-')    #finds the dash position
            lenrecord = len(teamrecord)     #gets the length of the record value
            ties = lenrecord - 1      #gets the position of the last dash
            equivwins = int(teamrecord[:dashpos]) + (0.5 * int(teamrecord[ties:]))      #wins + 1/2 * ties = equivalent wins
            #existingrecords.update( { teamname : equivwins} )    #updates the dictionary with the team/record pair
            standings_ws.cell(row=standingsrow, column=1).value = teamname
            standings_ws.cell(row=standingsrow, column=2).value = teamrecord
            standings_ws.cell(row=standingsrow, column=3).value = equivwins
            splitrecord = teamrecord.split('-')   #splits the record string using dash delimiter
            totalgamesplayed = int(splitrecord[0]) + int(splitrecord[1]) + int(splitrecord[2])      #sums wins losses and ties to get number of games played
            standingsrow += 1
        else:
            continue
    print("Standings Scrape Complete")
    standings_wb.save('standings.xlsx')

    time.sleep(5)
    driver.get(fantraxurl3)  
    time.sleep(5)

    #click the status/team dropdown
    status_selector = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/div/app-league-players/div/section/filter-panel/div/div[4]/div[1]/mat-form-field/div/div[1]/div[3]/mat-select/div/div[1]')
    driver.execute_script("arguments[0].click();", status_selector)

    driver.implicitly_wait(1)

    #select All Taken Players
    try:
        taken_selector = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div/div/mat-option[5]/span')
    except NoSuchElementException:
        taken_selector = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div/div/div/mat-option[5]/span')
    driver.execute_script("arguments[0].click();", taken_selector)

    time.sleep(3)  #gives time to load

    #select the rows per page dropdown
    perpage = driver.find_element_by_xpath('/html/body/app-root/div/div[2]/div/app-league-players/div/section/div[2]/pagination/div[4]/button/span')
    driver.execute_script("arguments[0].click();", perpage)

    driver.implicitly_wait(1)

    #select 500 per page
    try:
        fivehundredper = driver.find_element_by_xpath('/html/body/div[4]/div[2]/div/div/div/div/button[5]')
    except NoSuchElementException:
        fivehundredper = driver.find_element_by_xpath('/html/body/div[5]/div[2]/div/div/div/div/button[5]')
    driver.execute_script("arguments[0].click();", fivehundredper)

    time.sleep(3) #gives time to load

    soup_level2=BeautifulSoup(driver.page_source, 'lxml')

    writerow = 2
    
    #scrapes the player data table
    for tr in soup_level2.findAll("td", {"class": "ng-star-inserted"}):
        tds = tr.findAll("div", {"class": "scorer__info"})
        for td in tds:
            playername = td.findAll("div", {"class": "scorer__info__name"})
            playerpos = td.findAll("div", {"class": "scorer__info__positions"})
            for player in playername:
                fantrax_ws.cell(row=writerow, column=1).value = player.text     #writes the player name to the first column
            for player1 in playerpos:
                posandteam = player1.findAll("span")
                for entry in posandteam:
                    if len(entry.text) > 5 and entry.text.count(',') == 0:
                        continue
                    elif entry.text in ['C','1B','2B','3B','OF','SP','SS','RP','P','UT'] or entry.text.count(',') >= 1:
                        fantrax_ws.cell(row=writerow, column=2).value = entry.text        #writes the player position(s) to the 2nd column
                    elif entry.text == "(R)":
                        continue
                    elif entry.text == '-':
                        continue
                    elif entry.text == "":
                        continue
                    else:
                        fantrax_ws.cell(row=writerow, column=3).value = entry.text[1:]     #writes the player's MLB team to the 3rd column
            writerow += 1

    writerow = 2
    for tr1 in soup_level2.findAll("tr", {"class": "ng-star-inserted"}):
        tds1 = tr1.findAll("table-cell", {"class": "ng-star-inserted"})
        rowvalues = []
        for td1 in tds1:
            rowvalues.append(td1.text)
        fantrax_ws.cell(row=writerow, column=4).value = rowvalues[1]       #writes the player's fantasy team
        fantrax_ws.cell(row=writerow, column=5).value = rowvalues[5]       #writes the player's total points
        fantrax_ws.cell(row=writerow, column=6).value = rowvalues[6]       #writes the player's points per game
        writerow += 1

    
    
    print("Fantrax player scrape complete")
    fantrax_wb.save('fantrax.xlsx')


#### scrapes either the pitcher or hitter projections from fangraphs, specificy the worksheet to write to and the number of pages to scrape ####

def steamerscrape(link, pages, worksheet):
    driver = webdriver.Firefox()
    driver.implicitly_wait(8)
    driver.get(link)
    writerow = 2
    for page in range(1,pages):         
        rightbutton = driver.find_element_by_xpath("//button[@class='t-button rgActionButton rgPageNext']")
        soup_level1=BeautifulSoup(driver.page_source, 'lxml')
        
        for tr in soup_level1.findAll("tr", {"class": "rgRow"}):   #the table is structured where each row has an alternating tr-class, the first loop pulls the odds, second loop pulls the even rows
            tds = tr.findAll("td", {"class": "grid_line_regular"})
            writecol=1
            for td in tds:
                if writecol < 4:
                    worksheet.cell(row=writerow, column=writecol).value = td.text     #writes the name and team values into spreadsheet
                    writecol += 1
                else:
                    worksheet.cell(row=writerow, column=writecol).value = float(td.text)     #writes the statistical values into spreadsheet
                    writecol += 1
            writerow += 1
            
        for tr in soup_level1("tr", {"class": "rgAltRow"}):
            tds = tr.findAll("td", {"class": "grid_line_regular"})
            writecol=1
            for td in tds:
                if writecol < 4:
                    worksheet.cell(row=writerow, column=writecol).value = td.text     #writes the name and team values into spreadsheet
                    writecol += 1
                else:
                    worksheet.cell(row=writerow, column=writecol).value = float(td.text)     #writes the statistical values into spreadsheet
                    writecol += 1
            writerow += 1

        driver.execute_script("arguments[0].click();", rightbutton)       #clicks the right arrow to advance to the next page
        time.sleep(3)
    wb_write.save('steamer.xlsx')
    print("Scrape complete for: " + str(link))


#### converts each team name to the 2 or 3 letter acronym that matches Fantrax convention ####

def teamacronyms(worksheet):
    rows = worksheet.max_row
    for row in range(1,rows+1):
        for team in range(1,31):
            if worksheet.cell(row=row, column=3).value == ws4.cell(row=team, column=1).value:
                worksheet.cell(row=row, column=2).value = ws4.cell(row=team, column=2).value
            else:
                continue
    wb_write.save('steamer.xlsx')
    print("Acronyms assigned for: " + str(worksheet))


#### calculates the points per game for hitters ####
 
def hitterppgproj():
    hitterrows = ws_write.max_row
    for row in range(2,hitterrows+1):
        mobg = (ws_write.cell(row=row, column = 18).value * ws_write.cell(row=row, column = 4).value * 1.9651 - 36.2)         #MOBG
        hits = ws_write.cell(row=row, column = 7).value
        doubles = ws_write.cell(row=row, column = 8).value
        triples = ws_write.cell(row=row, column = 9).value
        hr = ws_write.cell(row=row, column = 10).value
        singles = hits - doubles - triples - hr
        run = ws_write.cell(row=row, column = 11).value
        rbi = ws_write.cell(row=row, column = 12).value
        bbhbp = ws_write.cell(row=row, column = 13).value + ws_write.cell(row=row, column = 15).value
        so = ws_write.cell(row=row, column = 14).value
        sb = ws_write.cell(row=row, column = 16).value
        cs = ws_write.cell(row=row, column = 17).value
        fpts = singles + doubles*2 + triples*3 + hr*4 + run*2 + rbi + bbhbp - so + sb*2 - cs*2 +mobg*2
        ppg = fpts / ws_write.cell(row=row, column = 4).value
        ws_write.cell(row=row, column = 26).value = fpts                      #Fantasy Points
        ws_write.cell(row=row, column = 27).value = ppg                       #PPG
    wb_write.save('steamer.xlsx')
    print("Hitter calculations complete")

#### calculates the points per game for pitchers ####

def pitcherppgproj():
    pitcherrows = ws2_write.max_row
    for row in range(2,pitcherrows+1):
        gs = ws2_write.cell(row=row, column=7).value
        pitcher_wins = ws2_write.cell(row=row, column=4).value
        pitcher_losses = ws2_write.cell(row=row, column=5).value
        ip = ws2_write.cell(row=row, column=10).value
        hitsallowed = ws2_write.cell(row=row, column=11).value
        er = ws2_write.cell(row=row, column=12).value
        pitcher_so = ws2_write.cell(row=row, column=14).value
        pitcher_bb = ws2_write.cell(row=row, column=15).value
        pitcher_fpts = (pitcher_wins*3 - pitcher_losses*3 + ip*3 - hitsallowed*0.5 - er*2 + pitcher_so - pitcher_bb) * 1.09    #1.09 coefficient accounts for the things that steamer does not project (eg. SHO, CG, NH)
        pitcher_ppg = pitcher_fpts / ws2_write.cell(row=row, column=8).value
        if gs > 4:
            if pitcher_ppg > 8.5:
                ws2_write.cell(row=row, column=23).value = pitcher_ppg
            else:
                ws2_write.cell(row=row, column=23).value = 8.5
        else:
            ws2_write.cell(row=row, column=23).value = 0
    wb_write.save('steamer.xlsx')
    print("Pitcher calculations complete")
        

####  links points per game projection from fangraphs steamer projection into the fantrax ownership sheet ####

def ppglinker():
    hitterrows = ws_write.max_row
    pitcherrows = ws2_write.max_row
    fantraxrows = fantrax_ws.max_row
    for player in range(2,fantraxrows):
        namelist = []
        namelist = fantrax_ws.cell(row=player, column=1).value.split()
        fantraxmlbteam = fantrax_ws.cell(row=player, column=3).value
        fantraxplayername = fantrax_ws.cell(row=player, column=1).value
        fantraxposition = fantrax_ws.cell(row=player, column=2).value
        if fantraxposition in ['SP','RP','SP,RP']:
            for fangraphplayer in range(2,pitcherrows): #links pitchers first
                fgnamelist = []
                fgnamelist = ws2_write.cell(row=fangraphplayer, column=1).value.split() 
                fgmlbteam = ws2_write.cell(row=fangraphplayer, column=2).value
                fgplayername = ws2_write.cell(row=fangraphplayer, column=1).value
                if namelist[0][:3].lower() == fgnamelist[0][:3].lower() and namelist[1][:4].lower() == fgnamelist[1][:4].lower() and fantraxmlbteam == fgmlbteam:
                    fantrax_ws.cell(row=player, column=7).value = ws2_write.cell(row=fangraphplayer, column=23).value
                elif fgplayername[:3].lower() == fantraxplayername[:3].lower() and fgplayername[-3:].lower() == fantraxplayername[-3:].lower() and fantraxmlbteam == fgmlbteam:
                    fantrax_ws.cell(row=player, column=7).value = ws2_write.cell(row=fangraphplayer, column=23).value
                else:
                    continue
        else:
            for fangraphplayer in range(2,hitterrows): #links hitters seconds
                fgnamelist = []
                fgnamelist = ws_write.cell(row=fangraphplayer, column=1).value.split() 
                fgmlbteam = ws_write.cell(row=fangraphplayer, column=2).value
                fgplayername = ws_write.cell(row=fangraphplayer, column=1).value
                if namelist[0][:3].lower() == fgnamelist[0][:3].lower() and namelist[1][:4].lower() == fgnamelist[1][:4].lower() and fantraxmlbteam == fgmlbteam:
                    fantrax_ws.cell(row=player, column=7).value = ws_write.cell(row=fangraphplayer, column=27).value
                elif fgplayername[:3].lower() == fantraxplayername[:3].lower() and fgplayername[-3:].lower() == fantraxplayername[-3:].lower() and fantraxmlbteam == fgmlbteam:
                    fantrax_ws.cell(row=player, column=7).value = ws_write.cell(row=fangraphplayer, column=27).value
                else: 
                    continue
    fantrax_wb.save('fantrax.xlsx')
    print("Linker complete")

#### Sorts the players from the fantrax spreadsheet to the optimal starting lineup and pitching rotation for each team ####    
    
def sortplayers():
    warnings.filterwarnings("ignore")
    fantraxdata = pd.read_excel('fantrax.xlsx')
    ownership = fantraxdata[["Player Name", "Position", "MLB Team", "Fantrax Team", "PPG"]]       #removes extraneous columns from the fantrax spreadsheet
    teams = ['AMH', 'BRK', 'CHI', 'DEN', 'FW', 'HAL', 'LA', 'NO', 'NP', 'PRI', 'TOR', 'WSH']       #teams in the fantasy league to loop through
    teamppg = []
    teambatrosters = {}
    batrosters_df = pd.DataFrame(columns = ["Team", "Team PPG"])
    print(batrosters_df)

    positionorder = ['C', '2B', 'SS', '3B', '1B', 'OF']    #priority order for assigning players to positions with the most scarce positions first
    
    for team in teams:
        replacementlevel = pd.DataFrame([["Replacement Level", "REPLC", "REPLC", team, 2.5]], columns=["Player Name", "Position", "MLB Team", "Fantrax Team", "PPG"])
        bestroster = ownership[ownership.PPG > 30]         #basically just creates a blank list by using a filter that applies to no players
        
        
        removeposlist = []                                  #blank list that will later be used to store the positions of the players to remove
        removenamelist = []                                 #blank list that will later be used to store the names of the players to remove
        for pos in positionorder:
            maxscorer = ownership[ownership.PPG == ownership.PPG.max()]         #pretty sure this isn't needed - confirm later
            teamroster2 = ownership[(ownership["Fantrax Team"] == team)]        #gets the complete roster of the specific team in the loop
            teamroster = teamroster2.fillna(0)                                  #converts the NAs to 0s
            specificteam = ownership[(ownership["Fantrax Team"] == team) & (ownership['Position'].str.contains(pos))]    #gets the list of players eligible at each position in the loop for each team in the loop
            if pos != 'OF':     #outfielders are seperated since we are selecting 3 OF and only 1 of each other position
                positionscorer = specificteam[specificteam.PPG == specificteam.PPG.max()]   #selects the player at that position group who is highest scorer
                positionscorer['Assigned Position'] = pos    #assigns the position in the loop to that player, since they may be eligible at multiple
                bestroster = pd.concat([positionscorer,bestroster])     #adds this player to the team's "best roster"
            else:
                positionscorer = specificteam.nlargest(3,'PPG')     #selects the three highest scoring players with OF eligibility
                positionscorer['Assigned Position'] = pos           #assigns these three players as OF
                bestroster = pd.concat([positionscorer,bestroster]) #adds these three players to best roster
        duplicates = bestroster[bestroster.duplicated(['Player Name'],keep=False)]      #finds the player(s) that appear twice on the best roster list (ie. they are eligible at more than one position), lists them twice, once for each position they are assigned to
        lengthdupes = len(duplicates.index)     #gets the length of this duplicate list 
        
        if lengthdupes > 0:          #only works on the teams that have a player assigned twice to different positions
            duplicateadd = bestroster[bestroster.duplicated(["Player Name"])]        #the list of duplicate players with each player appearing only once, we'll use this at the end
            replacementppg = []       #blank list that we'll be adding the ppg value for our selected replacement player
            replacementppg2 = []      #blank list that we'll be using for teams who have two duplicate players
            bestoptionnames = []      #blank list that we'll be using to add the names of the best option at each position where we need to find a replacement for a duplicate
            dupepos = duplicates['Assigned Position'].tolist()      #converts the duplicates in the dataframe to just a list of the duplicate positions
            
            for possearch in range(0,lengthdupes):      #loops through each duplicate player
                #replacement pool looks for all the players in the teamroster dataframe whose position matches the assigned position of the duplicate player position that is currently in the loop (possearch) & who do not already appear in the bestroster
                replacementpool = teamroster[(teamroster['Position'].str.contains(duplicates.iloc[possearch]['Assigned Position'])) & (~teamroster.index.isin(bestroster.index))]

                if len(replacementpool.index) > 0:      #makes sure there is a valid replacement option
                    bestoption = replacementpool[replacementpool.PPG == replacementpool.PPG.max()]
                    bestoptionnames.append(bestoption.iloc[0]["Player Name"])      #adds the players name to the bestoptionnames list
                else:       #skips the best option selection if there is no replacement pool at that position
                    continue
                
                #the duplicates list will always show a player two times, once for each position they are assigned, so for the first and second items in list, it goes to the first list and 3rd and 4th go to the second list
                if lengthdupes != 3:
                    if possearch in (0,1) and len(bestoption.index) > 0:   
                        replacementppg.append(bestoption.iloc[0]['PPG'])        #adds the ppg of the best options to the first list     
                    elif possearch in (2,3) and len(bestoption.index) > 0:
                        replacementppg2.append(bestoption.iloc[0]['PPG'])       #if there are more than 1 duplicate, adds ppg of the best options to second list
                    else:
                        continue
                else:
                    replacementppg.append(bestoption.iloc[0]['PPG'])        #adds the ppg of the best options to the first list
                    #replacementppg2.append(secondbestoption.iloc[0]['PPG'])        #adds the ppg of the best options to the first list
                    
            if len(replacementppg) > 1:                 #if there are best options at each position, select the max
                maxppgchoice = max(replacementppg)
            else:                                       #if there is only one best option, select it
                maxppgchoice = replacementppg[0]

            #replacement choice uses the ppg found in the last step to select the player, prevents mismatch by looking for the name in bestoptionnames
            replacementchoice = teamroster[(teamroster['PPG'] == maxppgchoice) & (teamroster['Player Name'].isin(bestoptionnames))]
            
           
            #same as above, determines the replacement for the second duplicate player
            if lengthdupes > 2:
                if len(replacementppg2) > 1:
                    maxppgchoice2 = max(replacementppg2)
                else:
                    maxppgchoice2 = replacementppg2[0]
                    
                replacementchoice2 = teamroster[(teamroster['PPG'] == maxppgchoice2) & (teamroster['Player Name'].isin(bestoptionnames))]
            else:
                continue
           

                        
            #posremoves selects the best roster minus the duplicate players
            posremoves = bestroster[(~bestroster.index.isin(duplicates.index))]

            
            #assembles the final roster by adding the barebones roster, with the removed duplicates and the replacement choice
            if lengthdupes < 2:    #for teams with only 1 duplicate player
                almostfinalroster = pd.concat([posremoves, duplicateadd, replacementchoice],axis=0,ignore_index=True)
            else:                   #for teams with two duplicate players
                almostfinalroster = pd.concat([posremoves, duplicateadd, replacementchoice, replacementchoice2],axis=0,ignore_index=True)

        else:
            almostfinalroster = bestroster    #if a team has no duplicates then it just converts the bestroster to the finalroster

        if team == 'WSH':
            print(replacementchoice)
            print(teamroster)
            print(duplicates)
            print(bestroster)
            print(almostfinalroster)
        hitters = teamroster[(teamroster["Position"] != "SP") & (teamroster["Position"] != "RP") & ~teamroster["Player Name"].isin(almostfinalroster["Player Name"])]  #gets the hitters not on the almostfinalroster
        utilityplayer = hitters[hitters.PPG == hitters.PPG.max()]   #gets the top ranked hitter from that list, selects as UTIL bat
        finalroster = pd.concat([almostfinalroster, utilityplayer],axis=0,ignore_index=True)   #adds the utility hitter 
        finalroster = finalroster[finalroster.PPG != 0]   #removes any people with zero
        

        while len(finalroster.index) <9:  #adds a replacement level player for each blank spot
            finalroster = pd.concat([finalroster, replacementlevel])

        print(finalroster[["Player Name", "Position", "Fantrax Team", "PPG"]])
        teamppg.append(finalroster["PPG"].sum())

    for i in range(0,len(teams)):    #updates a dictionary with the team name and their bats ppg
        teambatrosters.update( {teams[i]: teamppg[i]} )

    count = 1
    for item in teambatrosters.items():    #gets the dictionary created above into a dataframe
        batrosters_df.loc[count] = (item[0], item[1])
        count += 1
    print(batrosters_df.sort_values(by='Team PPG',ascending=False).round(2))
    

    
        

#### section where you call the various functions depending on what you need ####

##fantraxscrape()

##steamerscrape(url, 25, ws_write)               #scrapes hitter projections
##steamerscrape(url2, 25, ws2_write)              #scrapers pitcher projections
##
##teamacronyms(ws_write)                          #assign team acronyms to hitter sheet
##teamacronyms(ws2_write)                         #assign team acronyms to pitcher sheet
##
##hitterppgproj()                                #calculate hitter ppg
##pitcherppgproj()                               #calculate pitcher ppg
##

##ppglinker()                                    #link steamer projections into the fantrax sheet

sortplayers()                                   #sorts player rosters to find optimal starting lineup
